"""XLSX report generation for the fMRI Experimental Design Planner.

The workbook is the deliverable a human planner hands to a scanner technologist
or drops into an IRB / grant appendix: the solved design, the timing tables for
every aim, the session timeline, the efficiency diagnostics and the verbatim
scanner parameter cards that the design was solved against.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .protocols import PROTOCOL_ROLES

# ------------------------------------------------------------- WSU palette
GREEN = "046A38"
GOLD = "CBA052"
DARK_GREEN = "00482B"
DEEP_GOLD = "AE8643"
YELLOW = "F8E08E"
LEAF = "719949"
BEIGE = "DCD59A"
BLACK = "101820"
FADED_GOLD = "E7E3C6"
OFF_WHITE = "F2F1F0"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=OFF_WHITE)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=10, color=BLACK)
BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=BLACK)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color=DARK_GREEN)

TITLE_FILL = PatternFill("solid", fgColor=DARK_GREEN)
SECTION_FILL = PatternFill("solid", fgColor=GREEN)
HEADER_FILL = PatternFill("solid", fgColor=GREEN)
SUBHEADER_FILL = PatternFill("solid", fgColor=LEAF)
BAND_FILL = PatternFill("solid", fgColor=OFF_WHITE)
ALT_FILL = PatternFill("solid", fgColor=FADED_GOLD)
ACCENT_FILL = PatternFill("solid", fgColor=YELLOW)
WARN_FILL = PatternFill("solid", fgColor=BEIGE)

THIN = Side(style="thin", color=BEIGE)
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(vertical="top", wrap_text=True)
LEFT = Alignment(vertical="center", horizontal="left")
CENTER = Alignment(vertical="center", horizontal="center")
RIGHT = Alignment(vertical="center", horizontal="right")


# --------------------------------------------------------------- utilities


def _num(value: Any, default: float = 0.0) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return default
    if result != result or result in (float("inf"), float("-inf")):
        return default
    return result


def _fmt(value: float, digits: int = 1) -> str:
    return f"{_num(value):,.{digits}f}"


def _mmss(seconds: float) -> str:
    seconds = max(0.0, _num(seconds))
    minutes = int(seconds // 60)
    return f"{minutes:d}:{seconds - minutes * 60:04.1f}"


def _title_block(ws, title: str, subtitle: str, span: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = SUBTITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16
    return 4


def _section(ws, row: int, text: str, span: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1


def _table(
    ws,
    row: int,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    widths: Sequence[int] | None = None,
    wrap_columns: Sequence[int] = (),
    number_formats: Dict[int, str] | None = None,
) -> int:
    number_formats = number_formats or {}
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = GRID
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[row].height = 18
    row += 1

    for band, record in enumerate(rows):
        for index, value in enumerate(record, start=1):
            cell = ws.cell(row=row, column=index, value=value)
            cell.font = BODY_FONT
            cell.fill = BAND_FILL if band % 2 == 0 else ALT_FILL
            cell.border = GRID
            cell.alignment = WRAP if index in wrap_columns else LEFT
            if index in number_formats:
                cell.number_format = number_formats[index]
        row += 1

    if widths:
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
    return row + 1


def _kv_rows(ws, row: int, pairs: Sequence[Sequence[Any]], label_width: int = 38,
             value_width: int = 30) -> int:
    for band, (label, value) in enumerate(pairs):
        left = ws.cell(row=row, column=1, value=label)
        left.font = BODY_BOLD
        left.fill = BAND_FILL if band % 2 == 0 else ALT_FILL
        left.border = GRID
        left.alignment = LEFT
        right = ws.cell(row=row, column=2, value=value)
        right.font = BODY_FONT
        right.fill = BAND_FILL if band % 2 == 0 else ALT_FILL
        right.border = GRID
        right.alignment = LEFT
        row += 1
    ws.column_dimensions["A"].width = label_width
    ws.column_dimensions["B"].width = value_width
    return row + 1


# ------------------------------------------------------------------ sheets


def _sheet_summary(wb: Workbook, payload: Dict[str, Any]) -> None:
    meta = payload.get("meta", {})
    totals = payload.get("totals", {})
    budget = payload.get("budget", {})
    session = payload.get("session", {})
    aims = payload.get("aims", [])

    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = GREEN
    row = _title_block(
        ws,
        meta.get("studyTitle") or "fMRI Experimental Design",
        f"{meta.get('institution', 'Wright State University')}  |  generated "
        f"{payload.get('generated') or datetime.now().strftime('%Y-%m-%d %H:%M')}",
        span=6,
    )

    row = _section(ws, row, "Study", 6)
    row = _kv_rows(
        ws,
        row,
        [
            ("Study title", meta.get("studyTitle", "")),
            ("Principal investigator", meta.get("investigator", "")),
            ("Institution", meta.get("institution", "")),
            ("Participant identifier", meta.get("participantId", "")),
            ("Design identifier", meta.get("designId", "")),
            ("Notes", meta.get("notes", "")),
        ],
    )

    row = _section(ws, row, "Questions recorded by modality", 6)
    by_aim = totals.get("byAim", [])
    rows = [
        [
            record.get("name", ""),
            record.get("objective", ""),
            int(_num(record.get("primaryQuestions"))),
            int(_num(record.get("controlTrials"))),
            int(_num(record.get("trials"))),
            int(_num(record.get("questionsPerSession"))),
            int(_num(record.get("sessions"))),
            round(_num(record.get("hours")), 1),
        ]
        for record in by_aim
    ]
    rows.append(
        [
            "Total",
            "",
            int(_num(totals.get("primaryQuestions"))),
            int(_num(totals.get("controlTrials"))),
            int(_num(totals.get("trials"))),
            "",
            int(_num(totals.get("sessions"))),
            round(_num(totals.get("hours")), 1),
        ]
    )
    row = _table(
        ws,
        row,
        [
            "Modality",
            "Objective",
            "Questions recorded",
            "Control trials",
            "Total trials",
            "Questions / session",
            "Sessions",
            "Hours",
        ],
        rows,
        widths=[30, 16, 20, 15, 14, 19, 11, 10],
    )

    row = _section(ws, row, "Solved totals", 6)
    row = _kv_rows(
        ws,
        row,
        [
            ("Scanner-hour budget", f"{_fmt(budget.get('totalScannerHours'), 2)} h"),
            ("Scanner hours committed", f"{_fmt(totals.get('hours'), 2)} h"),
            ("Contingency reserve", f"{_fmt(budget.get('contingencyPct'), 0)} %"),
            ("Budget utilisation", f"{_fmt(totals.get('utilisationPct'), 1)} %"),
            ("Sessions planned", int(_num(totals.get("sessions")))),
            ("Session model", budget.get("sessionModel", "")),
            ("Solve mode", budget.get("solveMode", "")),
            ("Mean session duration", f"{_fmt(session.get('sessionMeanMinutes'))} min"),
            (
                "Session duration range",
                f"{_fmt(session.get('sessionMinMinutes'))} - "
                f"{_fmt(session.get('sessionMaxMinutes'))} min",
            ),
            ("Questions recorded", int(_num(totals.get("primaryQuestions")))),
            ("Control trials", int(_num(totals.get("controlTrials")))),
            ("Total trials", int(_num(totals.get("trials")))),
            (
                "Question goal",
                int(_num(totals.get("targetQuestions", budget.get("targetQuestionsTotal")))),
            ),
            ("Estimated raw data volume", f"{_fmt(totals.get('dataGB'), 1)} GB"),
            ("Calendar span", f"{_fmt(totals.get('weeks'), 1)} weeks"),
        ],
    )

    row = _section(ws, row, "Allocation by specific aim", 6)
    rows = []
    for aim in payload.get("aims", []):
        derived = aim.get("derived", {})
        rows.append(
            [
                aim.get("name", ""),
                aim.get("decode", {}).get("objective", ""),
                aim.get("protocolLabel", aim.get("protocol", "")),
                f"{_fmt(aim.get('requestedPct'), 1)} %",
                f"{_fmt(derived.get('sharePct'), 1)} %",
                f"{_fmt(derived.get('functionalHours'), 2)} h",
                int(_num(derived.get("totalTrials"))),
                int(_num(derived.get("primaryQuestions"))),
            ]
        )
    row = _table(
        ws,
        row,
        ["Aim", "Objective", "Bound protocol", "Requested share", "Realised share",
         "Functional hours", "Trials", "Questions"],
        rows,
        widths=[30, 16, 30, 16, 16, 18, 12, 12],
    )

    warnings = payload.get("warnings", [])
    row = _section(ws, row, "Constraint report", 6)
    if warnings:
        row = _table(
            ws,
            row,
            ["#", "Constraint message"],
            [[index, text] for index, text in enumerate(warnings, start=1)],
            widths=[6, 110],
            wrap_columns=(2,),
        )
    else:
        cell = ws.cell(row=row, column=1, value="All constraints satisfied; no clamps applied.")
        cell.font = NOTE_FONT
        row += 2


def _sheet_design_matrix(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Design Matrix")
    ws.sheet_properties.tabColor = GOLD
    row = _title_block(
        ws,
        "Hierarchical Design Matrix",
        "Trial to experiment expansion for every specific aim, solved against the current budget.",
        span=4,
    )
    for aim in payload.get("aims", []):
        name = aim.get("name", "")
        label = aim.get("protocolLabel", "")
        heading = name if label in ("", name) else f"{name} - {label}"
        heading += f"   [{aim.get('protocol', '')}.json]"
        row = _section(ws, row, heading, 4)
        rows = [
            [
                record.get("level", ""),
                record.get("sequence", ""),
                record.get("trials", ""),
                record.get("duration", ""),
            ]
            for record in aim.get("table", [])
        ]
        row = _table(
            ws,
            row,
            ["Level", "Sequence", "Trials", "Approximate duration"],
            rows,
            widths=[16, 86, 12, 26],
            wrap_columns=(2,),
        )


def _sheet_trial_structure(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Trial Structure")
    ws.sheet_properties.tabColor = LEAF
    row = _title_block(
        ws,
        "Trial Phase Structure",
        "Per-aim event sequence with jitter bounds and the expected value used for budgeting.",
        span=7,
    )
    for aim in payload.get("aims", []):
        derived = aim.get("derived", {})
        structure = aim.get("structure", {})
        row = _section(ws, row, f"{aim.get('name', '')} - trial phases", 7)
        rows = []
        for index, phase in enumerate(aim.get("phases", []), start=1):
            rows.append(
                [
                    index,
                    phase.get("name", ""),
                    _num(phase.get("min")),
                    _num(phase.get("max")),
                    round((_num(phase.get("min")) + _num(phase.get("max"))) / 2.0, 3),
                    "Yes" if phase.get("jitter") else "No",
                    phase.get("role", ""),
                ]
            )
        rows.append(
            [
                "",
                "TRIAL TOTAL",
                round(_num(derived.get("trialMin")), 2),
                round(_num(derived.get("trialMax")), 2),
                round(_num(derived.get("trialMean")), 2),
                "",
                "",
            ]
        )
        row = _table(
            ws,
            row,
            ["#", "Phase", "Min (s)", "Max (s)", "Expected (s)", "Jittered", "Model role"],
            rows,
            widths=[6, 26, 12, 12, 14, 12, 22],
            number_formats={3: "0.00", 4: "0.00", 5: "0.00"},
        )

        row = _section(ws, row, f"{aim.get('name', '')} - assembly", 7)
        decode = aim.get("decode", {})
        row = _kv_rows(
            ws,
            row,
            [
                ("Decoding objective", decode.get("objective", "")),
                ("Label ordering", decode.get("labelOrder", "")),
                ("Same-label run length", int(_num(decode.get("labelRunLength", 1)))),
                ("Trials per block", int(_num(structure.get("trialsPerBlock")))),
                ("Inter-trial gap (s)", _num(structure.get("interTrialGap"))),
                ("Blocks per run", int(_num(structure.get("blocksPerRun")))),
                ("Inter-block rest (s)", _num(structure.get("interBlockRest"))),
                ("Dummy volumes", int(_num(structure.get("dummyVolumes")))),
                ("Dummy duration (s)", round(_num(derived.get("dummySeconds")), 2)),
                ("Lead-in (s)", _num(structure.get("leadIn"))),
                ("Lead-out (s)", _num(structure.get("leadOut"))),
                ("Runs per session", int(_num(structure.get("runsPerSession")))),
                ("Sessions", int(_num(derived.get("sessions")))),
                ("TR (ms)", _num(aim.get("trMs"))),
                ("Volumes per run (dynamics)", int(_num(derived.get("volumesPerRun")))),
                ("Trials per run", int(_num(derived.get("trialsPerRun")))),
                ("Trials per session", int(_num(derived.get("trialsPerSession")))),
            ],
        )


def _sheet_budget(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Budget")
    ws.sheet_properties.tabColor = DEEP_GOLD
    budget = payload.get("budget", {})
    totals = payload.get("totals", {})
    caps = payload.get("caps", {})
    row = _title_block(
        ws,
        "Scanner-Hour Budget and Allocation",
        "Requested allocation, realised allocation and the constraint envelope used by the solver.",
        span=7,
    )

    row = _section(ws, row, "Budget envelope", 7)
    row = _kv_rows(
        ws,
        row,
        [
            ("Total scanner hours", _num(budget.get("totalScannerHours"))),
            ("Contingency reserve (%)", _num(budget.get("contingencyPct"))),
            ("Usable hours", _num(budget.get("usableHours"))),
            ("Hours committed", _num(totals.get("hours"))),
            ("Hours remaining", _num(totals.get("hoursRemaining"))),
            ("Solve mode", budget.get("solveMode", "")),
            ("Session model", budget.get("sessionModel", "")),
            ("Total question goal", int(_num(budget.get("targetQuestionsTotal")))),
            ("Sessions per week", _num(budget.get("sessionsPerWeek"))),
            ("Calendar weeks available", _num(budget.get("weeksAvailable"))),
        ],
    )

    row = _section(ws, row, "Allocation", 7)
    rows = []
    for aim in payload.get("aims", []):
        derived = aim.get("derived", {})
        rows.append(
            [
                aim.get("name", ""),
                _num(aim.get("requestedPct")),
                _num(derived.get("sharePct")),
                _num(derived.get("functionalHours")),
                _num(derived.get("overheadHours")),
                _num(derived.get("totalHours")),
                int(_num(derived.get("totalTrials"))),
                int(_num(derived.get("primaryQuestions"))),
            ]
        )
    row = _table(
        ws,
        row,
        [
            "Aim",
            "Requested %",
            "Realised %",
            "Functional h",
            "Overhead share h",
            "Total h",
            "Trials",
            "Questions",
        ],
        rows,
        widths=[28, 14, 14, 16, 18, 14, 12, 12],
        number_formats={2: "0.0", 3: "0.0", 4: "0.00", 5: "0.00", 6: "0.00"},
    )

    row = _section(ws, row, "Constraint envelope", 7)
    row = _kv_rows(
        ws,
        row,
        [
            ("Max run duration (min)", _num(caps.get("maxRunMinutes"))),
            ("Max session duration (min)", _num(caps.get("maxSessionMinutes"))),
            ("Max runs per session", int(_num(caps.get("maxRunsPerSession")))),
            ("Max sessions in study", int(_num(caps.get("maxSessionsTotal")))),
            ("Max continuous scanning (min)", _num(caps.get("maxContinuousMinutes"))),
            ("Min questions per aim", int(_num(caps.get("minQuestionsPerAim")))),
        ],
    )


def _timeline_rows(records: Sequence[Dict[str, Any]]) -> list:
    return [
        [
            record.get("order", ""),
            record.get("item", ""),
            record.get("protocolLabel", record.get("protocol", "")),
            round(_num(record.get("minutes")), 2),
            round(_num(record.get("cumulative")), 2),
            record.get("category", ""),
        ]
        for record in records
    ]


def _sheet_session_timeline(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Session Timeline")
    ws.sheet_properties.tabColor = BEIGE
    session = payload.get("session", {})
    timelines = session.get("timelines") or []
    row = _title_block(
        ws,
        "Session Timelines",
        "Console-order schedule per aim, including structural, reference and functional series.",
        span=6,
    )

    headers = ["#", "Series", "Protocol card", "Duration (min)", "Cumulative (min)", "Category"]
    widths = [6, 42, 34, 16, 18, 16]

    if timelines:
        for record in timelines:
            label = record.get("name", record.get("short", "Session"))
            detail = (
                f"{int(_num(record.get('runsPerSession')))} runs, "
                f"{_fmt(record.get('meanMinutes'))} min, "
                f"{int(_num(record.get('sessions')))} sessions, "
                f"{'custom' if record.get('custom') else 'shared'} composition"
            )
            row = _section(ws, row, f"{label} - {detail}", 6)
            row = _table(
                ws,
                row,
                headers,
                _timeline_rows(record.get("rows", [])),
                widths=widths,
                number_formats={4: "0.00", 5: "0.00"},
            )
            row += 1
    else:
        row = _table(
            ws,
            row,
            headers,
            _timeline_rows(session.get("timeline", [])),
            widths=widths,
            number_formats={4: "0.00", 5: "0.00"},
        )

    row = _section(ws, row, "Shared session composition", 6)
    row = _kv_rows(
        ws,
        row,
        [
            ("Screening (min)", _num(session.get("screeningMinutes"))),
            ("Positioning (min)", _num(session.get("positioningMinutes"))),
            ("Practice (min)", _num(session.get("practiceMinutes"))),
            ("Inter-run break (min)", _num(session.get("breakMinutes"))),
            ("Structural / reference block (min)", _num(session.get("setupMinutes"))),
            ("Mean session duration (min)", _num(session.get("sessionMeanMinutes"))),
            ("Shortest session (min)", _num(session.get("sessionMinMinutes"))),
            ("Longest session (min)", _num(session.get("sessionMaxMinutes"))),
        ],
    )


def _sheet_efficiency(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Efficiency")
    ws.sheet_properties.tabColor = YELLOW
    row = _title_block(
        ws,
        "Design Efficiency Diagnostics",
        "Simulated single run per aim: HRF-convolved regressors, contrast efficiency and collinearity.",
        span=8,
    )
    rows = []
    for aim in payload.get("aims", []):
        efficiency = aim.get("efficiency", {})
        decode = aim.get("decode", {})
        rows.append(
            [
                aim.get("name", ""),
                decode.get("objective", ""),
                decode.get("labelOrder", ""),
                int(_num(efficiency.get("volumes"))),
                round(_num(efficiency.get("sustainPct")), 1),
                round(_num(efficiency.get("saturationIndex")), 2),
                round(_num(efficiency.get("singleTrialEff")), 4),
                round(_num(efficiency.get("carryoverPct")), 2),
                round(_num(efficiency.get("promptBleedPct")), 2),
                round(_num(efficiency.get("effAnswerVsBaseline")), 2),
                round(_num(efficiency.get("effYesVsNo")), 2),
                round(_num(efficiency.get("effQuestionVsAnswer")), 2),
                round(_num(efficiency.get("corrQuestionAnswer")), 3),
                round(_num(efficiency.get("maxVif")), 2),
            ]
        )
    row = _table(
        ws,
        row,
        [
            "Aim",
            "Objective",
            "Label order",
            "Volumes",
            "Duty cycle %",
            "Stacking gain",
            "Single-trial eff.",
            "Carryover %",
            "Prompt bleed %",
            "Eff. answer vs baseline",
            "Eff. yes vs no",
            "Eff. question vs answer",
            "r(question, answer)",
            "Max VIF",
        ],
        rows,
        widths=[26, 15, 18, 10, 12, 13, 15, 12, 13, 20, 15, 20, 18, 10],
    )
    note = ws.cell(
        row=row,
        column=1,
        value=(
            "Efficiency is 1 / (c' (X'X)^-1 c) for a unit-scaled contrast on the simulated run "
            "design matrix (double-gamma HRF, constant, linear and quadratic drift regressors). "
            "Single-trial efficiency is the reciprocal mean variance of least-squares-all trial "
            "betas, which is what linear MVPA is trained on. Duty cycle is the median predicted task "
            "signal as a percentage of its 95th percentile: a saturating detection design holds it "
            "high because the response never settles, while a separation design drives it to zero. "
            "Stacking gain is the peak predicted signal divided by the peak of one isolated trial. "
            "Carryover is the previous answer response still present at the next prompt onset, and "
            "prompt bleed is the question response still present in the answer window, both as a "
            "percentage of a single-event peak. Values are comparable between candidate designs for "
            "the same aim, not across aims with different TRs."
        ),
    )
    note.font = NOTE_FONT
    note.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=14)


def _sheet_question_bank(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Question Bank")
    ws.sheet_properties.tabColor = LEAF
    bank = payload.get("questionBank", {})
    row = _title_block(
        ws,
        "Question Bank and Trial Composition",
        "Item supply, label balance and control-trial composition implied by the solved design.",
        span=5,
    )
    row = _section(ws, row, "Bank parameters", 5)
    row = _kv_rows(
        ws,
        row,
        [
            ("Question bank size", int(_num(bank.get("size")))),
            ("Maximum repeats per item", int(_num(bank.get("maxRepeats")))),
            ("Item supply (bank x repeats)", int(_num(bank.get("supply")))),
            ("Trials demanded by design", int(_num(bank.get("demand")))),
            ("Supply headroom", int(_num(bank.get("headroom")))),
            ("Yes label share (%)", _num(bank.get("yesPct"))),
            ("Control-trial share (%)", _num(bank.get("controlPct"))),
            ("Primary trials", int(_num(bank.get("primaryTrials")))),
            ("Control trials", int(_num(bank.get("controlTrials")))),
            ("Mean presentations per item", round(_num(bank.get("meanRepeats")), 2)),
        ],
    )
    families = bank.get("families", [])
    if families:
        row = _section(ws, row, "Question families", 5)
        rows = [
            [
                family.get("name", ""),
                _num(family.get("pct")),
                int(_num(family.get("trials"))),
                int(_num(family.get("items"))),
                family.get("role", ""),
            ]
            for family in families
        ]
        row = _table(
            ws,
            row,
            ["Family", "Share (%)", "Trials", "Items required", "Design role"],
            rows,
            widths=[32, 12, 12, 16, 60],
            wrap_columns=(5,),
        )


def _sheet_data_volume(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Data Volume")
    ws.sheet_properties.tabColor = BEIGE
    row = _title_block(
        ws,
        "Reconstructed Data Volume Estimate",
        "int16 magnitude reconstruction, no compression; add roughly 30 percent for derivatives.",
        span=8,
    )
    rows = []
    for aim in payload.get("aims", []):
        volume = aim.get("dataVolume", {})
        rows.append(
            [
                aim.get("name", ""),
                volume.get("matrix", ""),
                int(_num(volume.get("slices"))),
                int(_num(volume.get("volumesPerRun"))),
                round(_num(volume.get("mbPerRun")), 1),
                round(_num(volume.get("gbPerSession")), 2),
                round(_num(volume.get("gbTotal")), 2),
                volume.get("voxel", ""),
            ]
        )
    _table(
        ws,
        row,
        [
            "Aim",
            "Recon matrix",
            "Slices",
            "Volumes/run",
            "MB per run",
            "GB per session",
            "GB total",
            "Voxel (mm)",
        ],
        rows,
        widths=[26, 16, 10, 14, 14, 16, 14, 20],
    )


def _sheet_methods(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Methods Text")
    ws.sheet_properties.tabColor = DARK_GREEN
    row = _title_block(
        ws,
        "Generated Methods Text",
        "Paste-ready narrative generated from the solved design; verify before submission.",
        span=1,
    )
    ws.column_dimensions["A"].width = 132
    for paragraph in str(payload.get("methodsText", "")).split("\n\n"):
        cell = ws.cell(row=row, column=1, value=paragraph.strip())
        cell.font = BODY_FONT
        cell.alignment = WRAP
        ws.row_dimensions[row].height = max(30, 14 * (len(paragraph) // 110 + 1))
        row += 2


def _sheet_markdown(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Markdown Tables")
    ws.sheet_properties.tabColor = GOLD
    row = _title_block(
        ws,
        "Markdown Export",
        "Design matrix tables in GitHub-flavoured Markdown, one row per line.",
        span=1,
    )
    ws.column_dimensions["A"].width = 150
    for name, markdown in (payload.get("markdownTables") or {}).items():
        cell = ws.cell(row=row, column=1, value=name)
        cell.font = BODY_BOLD
        cell.fill = ACCENT_FILL
        row += 1
        for line in str(markdown).split("\n"):
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = Font(name="Consolas", size=9, color=BLACK)
            row += 1
        row += 1


INVALID_SHEET_CHARS = r"[]:*?/\\"


def _sheet_name(wb: Workbook, preferred: str, fallback: str) -> str:
    """Excel rejects []:*?/\\ and titles longer than 31 characters."""
    cleaned = "".join(" " if ch in INVALID_SHEET_CHARS else ch for ch in (preferred or fallback))
    cleaned = " ".join(cleaned.split()).strip("'")[:28] or fallback[:28]
    name = cleaned
    for suffix in range(2, 40):
        if name not in wb.sheetnames:
            return name
        name = f"{cleaned[:26]} {suffix}"
    return fallback[:31]


def _sheet_protocol(wb: Workbook, slug: str, data: Dict[str, Any], bound_to: str) -> None:
    label = PROTOCOL_ROLES.get(slug, {}).get("label", slug)
    ws = wb.create_sheet(_sheet_name(wb, label, slug))
    ws.sheet_properties.tabColor = GREEN if bound_to else BEIGE
    subtitle = f"{slug}.json"
    if bound_to:
        subtitle += f"  |  bound to {bound_to}"
    row = _title_block(ws, label, subtitle, span=3)

    for section, rows in data.items():
        if section.startswith("_") or not isinstance(rows, list):
            continue
        row = _section(ws, row, section, 3)
        records = []
        for record in rows:
            indent = int(_num(record.get("indent")))
            records.append(
                [
                    ("    " * indent) + str(record.get("parameter", "")),
                    str(record.get("value", "")),
                    indent,
                ]
            )
        row = _table(
            ws,
            row,
            ["Parameter", "Value", "Indent"],
            records,
            widths=[44, 34, 9],
        )


# ------------------------------------------------------------------- entry


def build_workbook(payload: Dict[str, Any], protocols: Dict[str, Any]) -> bytes:
    """Render the full planner report and return the .xlsx bytes."""
    wb = Workbook()
    _sheet_summary(wb, payload)
    _sheet_design_matrix(wb, payload)
    _sheet_trial_structure(wb, payload)
    _sheet_budget(wb, payload)
    _sheet_session_timeline(wb, payload)
    _sheet_efficiency(wb, payload)
    _sheet_question_bank(wb, payload)
    _sheet_data_volume(wb, payload)
    _sheet_methods(wb, payload)
    _sheet_markdown(wb, payload)

    bindings = {}
    for aim in payload.get("aims", []):
        if aim.get("protocol"):
            bindings.setdefault(aim["protocol"], []).append(aim.get("name", ""))
    for record in payload.get("session", {}).get("timeline", []):
        if record.get("protocol") and record.get("category") != "Functional":
            bindings.setdefault(record["protocol"], []).append("Session block")

    for slug, data in protocols.items():
        if not isinstance(data, dict) or "_error" in data:
            continue
        bound = ", ".join(dict.fromkeys(bindings.get(slug, [])))
        _sheet_protocol(wb, slug, data, bound)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
